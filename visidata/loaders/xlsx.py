from visidata import *

vd.option('xlsx_cellobject', False, 'include column of cell objects', replay=True)
vd.option('xlsx_fontcolor', False, 'include cell font colors', replay=True)
vd.option('xlsx_fillcolor', False, 'include cell fill colors', replay=True)

@VisiData.api
def open_xls(vd, p):
    return XlsIndexSheet(p.name, source=p)

@VisiData.api
def open_xlsx(vd, p):
    return XlsxIndexSheet(p.name, source=p)

class XlsxIndexSheet(IndexSheet):
    'Load XLSX file (in Excel Open XML format).'
    rowtype = 'sheets'  # rowdef: xlsxSheet
    columns = [
        Column('sheet', getter=lambda col,row: row.source.title),  # xlsx sheet title
        ColumnAttr('name', width=0),  # visidata Sheet name
        ColumnAttr('nRows', type=int),
        ColumnAttr('nCols', type=int),
        Column('active', getter=lambda col,row: row.source is col.sheet.workbook.active),
    ]
    nKeys = 1

    def iterload(self):
        import openpyxl
        self.workbook = openpyxl.load_workbook(str(self.source), data_only=True, read_only=True)
        for sheetname in self.workbook.sheetnames:
            src = self.workbook[sheetname]
            yield XlsxSheet(self.name, sheetname, source=src)

class XlsxSheet(SequenceSheet):
    # rowdef: AttrDict of column_letter to cell
    def setCols(self, headerrows):
        from openpyxl.utils.cell import get_column_letter
        self.columns = []
        self._rowtype = AttrDict

        if not headerrows:
            return

        headers = [[cell.value for cell in row.values()] for row in headerrows]
        column_letters = [
                x.column_letter if 'column_letter' in dir(x)
                else get_column_letter(i+1)
                for i, x in enumerate(headerrows[0].values())]

        for i, colnamelines in enumerate(itertools.zip_longest(*headers, fillvalue='')):
            colnamelines = ['' if c is None else c for c in colnamelines]
            column_name = ''.join(map(str, colnamelines))
            self.addColumn(AttrColumn(column_name, column_letters[i] + '.value'))
            self.addXlsxColumns(column_letters[i], column_name)

    def addRow(self, row, index=None):
        Sheet.addRow(self, row, index=index)  # skip SequenceSheet
        for column_letter, v in list(row.items())[len(self.columns):len(row)]:  # no-op if already done
            self.addColumn(AttrColumn('', column_letter + '.value'))
            self.addXlsxColumns(column_letter, column_letter)

    def iterload(self):
        from openpyxl.utils.cell import get_column_letter
        worksheet = self.source
        for row in Progress(worksheet.iter_rows(), total=worksheet.max_row or 0):
            yield AttrDict({get_column_letter(i+1): cell for i, cell in enumerate(row)})

    def addXlsxColumns(self, column_letter, column_name):
        if self.options.xlsx_cellobject:
            self.addColumn(
                    AttrColumn(column_name + '_cellPyObj', column_letter))
        if self.options.xlsx_fontcolor:
            self.addColumn(
                    AttrColumn(column_name + '_fontColor',
                        column_letter + '.font.color.value'))
        if self.options.xlsx_fillcolor:
            self.addColumn(
                    AttrColumn(column_name + '_fillColor', column_letter +
                        '.fill.start_color.value'))

class XlsIndexSheet(IndexSheet):
    'Load XLS file (in Excel format).'
    rowtype = 'sheets'  # rowdef: xlsSheet
    columns = [
        Column('sheet', getter=lambda col,row: row.source.name),  # xls sheet name
        ColumnAttr('name', width=0),  # visidata sheet name
        ColumnAttr('nRows', type=int),
        ColumnAttr('nCols', type=int),
    ]
    nKeys = 1
    def iterload(self):
        import xlrd
        self.workbook = xlrd.open_workbook(str(self.source))
        for sheetname in self.workbook.sheet_names():
            yield XlsSheet(self.name, sheetname, source=self.workbook.sheet_by_name(sheetname))


class XlsSheet(SequenceSheet):
    def iterload(self):
        worksheet = self.source
        for rownum in Progress(range(worksheet.nrows)):
            yield list(worksheet.cell(rownum, colnum).value for colnum in range(worksheet.ncols))


def xls_name(name):
    # sheet name can not be longer than 31 characters
    xname = cleanName(name)[:31]
    if xname != name:
        vd.warning(f'{name} saved as {xname}')
    return xname


@VisiData.api
def save_xlsx(vd, p, *sheets):
    import openpyxl

    wb = openpyxl.Workbook()
    wb.remove_sheet(wb['Sheet'])

    for vs in sheets:
        ws = wb.create_sheet(title=xls_name(vs.name))

        headers = [col.name for col in vs.visibleCols]
        ws.append(headers)

        for dispvals in vs.iterdispvals(format=False):

            row = []
            for col, v in dispvals.items():
                if col.type == date:
                    v = datetime.datetime.fromtimestamp(int(v.timestamp()))
                elif not vd.isNumeric(col):
                    v = str(v)
                row.append(v)

            ws.append(row)

    wb.active = ws

    wb.save(filename=p)
    vd.status(f'{p} save finished')


@VisiData.api
def save_xls(vd, p, *sheets):
    import xlwt

    wb = xlwt.Workbook()

    for vs in sheets:
        ws1 = wb.add_sheet(xls_name(vs.name))
        for col_i, col in enumerate(vs.visibleCols):
            ws1.write(0, col_i, col.name)

        for r_i, dispvals in enumerate(vs.iterdispvals(format=True)):
            r_i += 1
            for c_i, v in enumerate(dispvals.values()):
                ws1.write(r_i, c_i, v)

    wb.save(p)
    vd.status(f'{p} save finished')
